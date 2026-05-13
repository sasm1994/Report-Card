import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "marks_database.xlsx")

# ════════════════════════════════════════════════════════════
# SUBJECT GROUPS — one per template type
# ════════════════════════════════════════════════════════════

SUBJECT_GROUPS = {

    # NUR-Template.html — Annual exam, marks out of 50
    "NUR": [
        "ENGLISH", "MATHEMATICS",
        "GENERAL KNOWLEDGE", "RHYMES AND STORY TELLING",
        # Co-Scholastic
        "ATTENDANCE_GRADE", "ART_GRADE", "DISCIPLINE_GRADE",
        # Attendance
        "DAYS_PRESENT", "DAYS_WORKING"
    ],

    # KG-Template.html — LKG / UKG, single semester
    "KG": [
        # Main subjects: PT, B/C, Project, SEM, Total, Grade
        "ENG_PT", "ENG_BC", "ENG_PROJECT", "ENG_SEM", "ENG_TOTAL", "ENG_GRADE",
        "MATH_PT", "MATH_BC", "MATH_PROJECT", "MATH_SEM", "MATH_TOTAL", "MATH_GRADE",
        "EVS_PT", "EVS_BC", "EVS_PROJECT", "EVS_SEM", "EVS_TOTAL", "EVS_GRADE",
        "HINDI_PT", "HINDI_BC", "HINDI_PROJECT", "HINDI_SEM", "HINDI_TOTAL", "HINDI_GRADE",
        # Grand total
        "GRAND_TOTAL", "PERCENTAGE",
        # Computer (Practical / Portfolio / Semester)
        "COMP_PRACTICAL", "COMP_PORTFOLIO", "COMP_SEM", "COMP_TOTAL", "COMP_GRADE",
        # Co-Scholastic
        "RHYMES_GRADE", "ART_GRADE", "HPE_GRADE", "DISCIPLINE_GRADE",
        # Attendance
        "DAYS_PRESENT", "DAYS_WORKING"
    ],

    # 1-2-Template.html — Class 1 & 2, single semester, adds ODIA
    "CLASS_1_2": [
        # Main subjects
        "ENG_PT", "ENG_BC", "ENG_PROJECT", "ENG_SEM", "ENG_TOTAL", "ENG_GRADE",
        "MATH_PT", "MATH_BC", "MATH_PROJECT", "MATH_SEM", "MATH_TOTAL", "MATH_GRADE",
        "EVS_PT", "EVS_BC", "EVS_PROJECT", "EVS_SEM", "EVS_TOTAL", "EVS_GRADE",
        "HINDI_PT", "HINDI_BC", "HINDI_PROJECT", "HINDI_SEM", "HINDI_TOTAL", "HINDI_GRADE",
        "ODIA_PT", "ODIA_BC", "ODIA_PROJECT", "ODIA_SEM", "ODIA_TOTAL", "ODIA_GRADE",
        # Grand total
        "GRAND_TOTAL", "PERCENTAGE",
        # Computer
        "COMP_PRACTICAL", "COMP_PORTFOLIO", "COMP_SEM", "COMP_TOTAL", "COMP_GRADE",
        # Co-Scholastic
        "HPE_GRADE", "ART_GRADE", "WORK_GRADE", "DISCIPLINE_GRADE",
        # Attendance
        "DAYS_PRESENT", "DAYS_WORKING"
    ],

    # 3-9-Template.html — Class 3 to 9, BOTH semesters
    "CLASS_3_9": [
        # ── Semester I ──
        "ENG_PT1", "ENG_MA1", "ENG_PF1", "ENG_SE1", "ENG_SEM1", "ENG_TOTAL1", "ENG_GRADE1",
        "MATH_PT1", "MATH_MA1", "MATH_PF1", "MATH_SE1", "MATH_SEM1", "MATH_TOTAL1", "MATH_GRADE1",
        "SCI_PT1", "SCI_MA1", "SCI_PF1", "SCI_SE1", "SCI_SEM1", "SCI_TOTAL1", "SCI_GRADE1",
        "SST_PT1", "SST_MA1", "SST_PF1", "SST_SE1", "SST_SEM1", "SST_TOTAL1", "SST_GRADE1",
        "HINDI_PT1", "HINDI_MA1", "HINDI_PF1", "HINDI_SE1", "HINDI_SEM1", "HINDI_TOTAL1", "HINDI_GRADE1",
        "ODIA_PT1", "ODIA_MA1", "ODIA_PF1", "ODIA_SE1", "ODIA_SEM1", "ODIA_TOTAL1", "ODIA_GRADE1",
        # Computer Sem I
        "COMP_PRAC1", "COMP_BC1", "COMP_SEM1", "COMP_TOTAL1", "COMP_GRADE1",
        # Co-Scholastic Sem I
        "HPE_GRADE1", "ART_GRADE1", "WORK_GRADE1", "DISCIPLINE_GRADE1",

        # ── Semester II ──
        "ENG_PT2", "ENG_MA2", "ENG_PF2", "ENG_SE2", "ENG_SEM2", "ENG_TOTAL2", "ENG_GRADE2",
        "MATH_PT2", "MATH_MA2", "MATH_PF2", "MATH_SE2", "MATH_SEM2", "MATH_TOTAL2", "MATH_GRADE2",
        "SCI_PT2", "SCI_MA2", "SCI_PF2", "SCI_SE2", "SCI_SEM2", "SCI_TOTAL2", "SCI_GRADE2",
        "SST_PT2", "SST_MA2", "SST_PF2", "SST_SE2", "SST_SEM2", "SST_TOTAL2", "SST_GRADE2",
        "HINDI_PT2", "HINDI_MA2", "HINDI_PF2", "HINDI_SE2", "HINDI_SEM2", "HINDI_TOTAL2", "HINDI_GRADE2",
        "ODIA_PT2", "ODIA_MA2", "ODIA_PF2", "ODIA_SE2", "ODIA_SEM2", "ODIA_TOTAL2", "ODIA_GRADE2",
        # Computer Sem II
        "COMP_PRAC2", "COMP_BC2", "COMP_SEM2", "COMP_TOTAL2", "COMP_GRADE2",
        # Co-Scholastic Sem II
        "HPE_GRADE2", "ART_GRADE2", "WORK_GRADE2", "DISCIPLINE_GRADE2",

        # ── Attendance & Extra ──
        "DAYS_PRESENT", "DAYS_WORKING",
        "ADMISSION_NO", "PROMOTED_TO"
    ]
}


# ════════════════════════════════════════════════════════════
# HELPER — map class sheet name → subject group key
# ════════════════════════════════════════════════════════════

def get_subject_group(class_name):
    """
    Returns the SUBJECT_GROUPS key for a given sheet name.
    Examples:
        NUR_A       → NUR
        LKG_B       → KG
        Class_1_A   → CLASS_1_2
        Class_3_B   → CLASS_3_9
    """
    name = class_name.upper()
    if "NUR" in name:
        return "NUR"
    elif "LKG" in name or "UKG" in name:
        return "KG"
    elif any(f"CLASS_{n}_" in name for n in ["1", "2"]):
        return "CLASS_1_2"
    else:
        return "CLASS_3_9"


def get_subjects(class_name):
    """Return the subject column list for a class."""
    return SUBJECT_GROUPS[get_subject_group(class_name)]


# ════════════════════════════════════════════════════════════
# MAIN DATABASE CLASS
# ════════════════════════════════════════════════════════════

class MarksDatabase:

    def __init__(self, filename=EXCEL_FILE):
        self.filename = filename
        self.ensure_file_exists()

    # ── Create blank database if not present ────────────────
    def ensure_file_exists(self):
        if not os.path.exists(self.filename):
            self.create_blank_database()

    def create_blank_database(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        classes = [
            # Pre-Primary
            "NUR_A", "NUR_B",
            "LKG_A", "LKG_B",
            "UKG_A", "UKG_B",
            # Primary
            "Class_1_A", "Class_1_B",
            "Class_2_A", "Class_2_B",
            "Class_3_A", "Class_3_B",
            "Class_4_A", "Class_4_B",
            "Class_5_A", "Class_5_B",
            # Middle
            "Class_6_A", "Class_6_B",
            "Class_7_A", "Class_7_B",
            "Class_8_A", "Class_8_B",
            # Secondary
            "Class_9_A", "Class_9_B",
            "Class_10_A", "Class_10_B",
        ]

        for cls in classes:
            ws = wb.create_sheet(title=cls)
            self.format_sheet(ws, cls)

        wb.save(self.filename)
        print(f"✓ Created blank database: {self.filename}")

    # ── Format a sheet with correct headers for its class ───
    def format_sheet(self, ws, class_name="Class_3_A"):
        subjects = get_subjects(class_name)
        headers  = ["StudentName", "RollNo"] + subjects

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value     = header
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill(start_color="4472C4",
                                         end_color="4472C4",
                                         fill_type="solid")
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)

        # Column widths
        ws.column_dimensions["A"].width = 28   # Student Name
        ws.column_dimensions["B"].width = 10   # Roll No
        for col in range(3, 3 + len(subjects)):
            ws.column_dimensions[get_column_letter(col)].width = 14

        ws.row_dimensions[1].height = 30

    # ── Get all class sheet names ────────────────────────────
    def get_all_classes(self):
        try:
            wb = openpyxl.load_workbook(self.filename)
            return wb.sheetnames
        except Exception as e:
            print(f"Error reading classes: {e}")
            return []

    # ── Get all students in a class ─────────────────────────
    def get_students(self, class_name):
        try:
            wb = openpyxl.load_workbook(self.filename)
            if class_name not in wb.sheetnames:
                return []

            ws       = wb[class_name]
            subjects = get_subjects(class_name)
            students = []

            for row_idx in range(2, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=1).value
                roll = ws.cell(row=row_idx, column=2).value

                if not name or not roll:
                    continue

                student = {
                    "StudentName" : str(name).strip(),
                    "RollNo"      : roll,
                    "row_idx"     : row_idx,
                    "group"       : get_subject_group(class_name)
                }

                # Read all subject columns
                for subj_idx, subject in enumerate(subjects, start=3):
                    val = ws.cell(row=row_idx, column=subj_idx).value
                    student[subject] = val if val is not None else ""

                students.append(student)

            return students

        except Exception as e:
            print(f"Error reading students from {class_name}: {e}")
            return []

    # ── Save marks for a student ─────────────────────────────
    def save_marks(self, class_name, roll_no, marks_dict):
        """
        marks_dict: { "ENG_PT1": 8, "ENG_MA1": 5, ... }
        """
        try:
            wb = openpyxl.load_workbook(self.filename)
            if class_name not in wb.sheetnames:
                return False, "Class not found"

            ws       = wb[class_name]
            subjects = get_subjects(class_name)

            # Find student row
            student_row = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=2).value == roll_no:
                    student_row = row_idx
                    break

            if not student_row:
                return False, "Student not found"

            # Update only columns that are in marks_dict
            for subj_idx, subject in enumerate(subjects, start=3):
                if subject in marks_dict:
                    ws.cell(row=student_row,
                            column=subj_idx).value = marks_dict[subject]

            wb.save(self.filename)
            return True, "Marks saved successfully"

        except Exception as e:
            return False, str(e)

    # ── Add a student to a class ─────────────────────────────
    def add_student(self, class_name, name, roll_no):
        try:
            wb = openpyxl.load_workbook(self.filename)
            if class_name not in wb.sheetnames:
                return False, f"Sheet '{class_name}' not found"

            ws      = wb[class_name]
            new_row = ws.max_row + 1

            ws.cell(row=new_row, column=1).value = str(name).strip().upper()
            ws.cell(row=new_row, column=2).value = int(roll_no)

            wb.save(self.filename)
            return True, "Student added"

        except Exception as e:
            return False, str(e)

    # ── Export class marks to CSV ────────────────────────────
    def export_to_csv(self, class_name, output_file):
        import csv
        try:
            students = self.get_students(class_name)
            subjects = get_subjects(class_name)
            fieldnames = ["StudentName", "RollNo"] + subjects

            with open(output_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames,
                                        extrasaction="ignore")
                writer.writeheader()
                for student in students:
                    writer.writerow(student)

            return True, f"Exported to {output_file}"

        except Exception as e:
            return False, str(e)

    # ── Get subject list for a class (used in API/templates) ─
    def get_subject_list(self, class_name):
        return get_subjects(class_name)