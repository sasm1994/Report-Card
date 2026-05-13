import openpyxl
import os
import glob

# ── CONFIG ───────────────────────────────────────────────────
# Folder where all your class Excel files are kept
# Put all files like 3-B_Students.xlsx, 4-A_Students.xlsx etc. in this folder
SOURCE_FOLDER = r"student_lists"

# Your marks database file

DB_FILE = r"marks_database.xlsx"
# ────────────────────────────────────────────────────────────

def parse_class_section(class_val, section_val):
    """
    Convert 'STD 3' + 'B'  →  'Class_3_B'
    Convert 'STD 10' + 'A' →  'Class_10_A'
    """
    class_val   = str(class_val).strip().upper()
    section_val = str(section_val).strip().upper()

    # Remove 'STD ' prefix
    class_val = class_val.replace("STD ", "").replace("CLASS ", "").strip()

    return f"Class_{class_val}_{section_val}"

def import_file(src_file, db_wb):
    print(f"\n📂 Processing: {os.path.basename(src_file)}")

    src_wb = openpyxl.load_workbook(src_file)
    src_ws = src_wb.active

    added   = 0
    skipped = 0

    # ── Auto-detect which row data starts from ──────────────
    # Look for the first row where Column A has a NUMBER (= Roll No)
    data_start_row = None
    for row in src_ws.iter_rows(min_row=1, values_only=True):
        if row[0] is not None:
            try:
                int(float(str(row[0])))   # Check if Column A is a number
                data_start_row = row      # This is the first data row
                break
            except (ValueError, TypeError):
                continue   # Skip title/header rows

    if data_start_row is None:
        print(f"  ⚠ No student data found in file — skipping")
        return 0

    # Find the actual row index of data_start_row
    start_idx = None
    for idx, row in enumerate(src_ws.iter_rows(min_row=1, values_only=True), start=1):
        if row == data_start_row:
            start_idx = idx
            break

    print(f"  ℹ Data starts at row: {start_idx}")
    # ────────────────────────────────────────────────────────

    for row in src_ws.iter_rows(min_row=start_idx, values_only=True):

        # Skip completely empty rows
        if not any(row):
            continue

        # Safely read each column
        roll    = row[0] if len(row) > 0 else None
        name    = row[1] if len(row) > 1 else None
        cls     = row[2] if len(row) > 2 else None
        section = row[3] if len(row) > 3 else None

        # Skip if name or roll is missing
        if not name or not roll:
            continue

        # Skip if roll is not a number (catches leftover header rows)
        try:
            roll = int(float(str(roll)))
        except (ValueError, TypeError):
            continue

        # ── If Section column is missing, extract from filename ──
        if not section:
            filename = os.path.basename(src_file)
            try:
                section = filename.split("-")[1].split("_")[0].strip().upper()
                print(f"  ℹ Section extracted from filename: '{section}'")
            except:
                print(f"  ⚠ Could not determine section for {name} — skipping")
                skipped += 1
                continue

        # ── If Class column is missing, extract from filename ──
        if not cls:
            filename = os.path.basename(src_file)
            try:
                cls = filename.split("-")[0].strip()
                print(f"  ℹ Class extracted from filename: '{cls}'")
            except:
                print(f"  ⚠ Could not determine class for {name} — skipping")
                skipped += 1
                continue

        # Build sheet name e.g. Class_6_A
        sheet_name = parse_class_section(cls, section)

        # Check if sheet exists in database
        if sheet_name not in db_wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' not found — skipping {name}")
            skipped += 1
            continue

        ws      = db_wb[sheet_name]
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1).value = str(name).strip().upper()
        ws.cell(row=new_row, column=2).value = roll

        print(f"  ✓  Roll {roll:>3} | {name} → {sheet_name}")
        added += 1

    print(f"  ── Added: {added} | Skipped: {skipped}")
    return added

def main():
    # Load database
    if not os.path.exists(DB_FILE):
        print(f"❌ Database not found: {DB_FILE}")
        return

    db_wb = openpyxl.load_workbook(DB_FILE)
    print(f"✅ Opened database: {DB_FILE}")
    print(f"   Sheets available: {db_wb.sheetnames}\n")

    # Find all Excel files in source folder
    all_files = glob.glob(os.path.join(SOURCE_FOLDER, "*.xlsx"))

    if not all_files:
        print(f"❌ No .xlsx files found in: {SOURCE_FOLDER}")
        return

    total_added = 0
    for f in all_files:
        total_added += import_file(f, db_wb)

    # Save database
    db_wb.save(DB_FILE)
    print(f"\n✅ DONE! Total students added: {total_added}")
    print(f"   Database saved: {DB_FILE}")

if __name__ == "__main__":
    main()